from fastapi import APIRouter, Depends, HTTPException, Header, UploadFile, File
from fastapi.responses import StreamingResponse
from io import BytesIO, StringIO
import csv
import uuid
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import get_db_connection
from auth import get_current_user
from routes.inventario import get_ente_id

router = APIRouter(prefix="/api/inventario", tags=["Inventario - Import/Export"])

COLONNE_CSV = [
    "Numero", "Descrizione", "Categoria", "Ubicazione", "Quantita",
    "Stato Conservazione", "Valore Stimato", "Valore Assicurato",
    "Data Acquisto", "Fornitore", "Provenienza", "Note"
]


# ============================================
# EXPORT CSV
# ============================================

@router.get("/export/csv")
async def export_csv(
    current_user: dict = Depends(get_current_user),
    x_ente_id: str = Header(None, alias="X-Ente-Id")
):
    ente_id = get_ente_id(current_user, x_ente_id)
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT
                b.numero_progressivo, b.descrizione, c.nome, u.nome,
                b.quantita, b.stato_conservazione, b.valore_stimato,
                b.valore_assicurato, b.data_acquisto, b.fornitore,
                b.provenienza, b.note
            FROM beni_inventario b
            LEFT JOIN inventario_categorie c ON b.categoria_id = c.id
            LEFT JOIN inventario_ubicazioni u ON b.ubicazione_id = u.id
            WHERE b.ente_id = %s AND b.stato = 'attivo'
            ORDER BY b.numero_progressivo
        """, (ente_id,))
        rows = cur.fetchall()

        output = StringIO()
        # BOM per Excel
        output.write('\ufeff')
        writer = csv.writer(output, delimiter=';')
        writer.writerow(COLONNE_CSV)
        for r in rows:
            writer.writerow([
                r[0],  # numero
                r[1] or '',  # descrizione
                r[2] or '',  # categoria
                r[3] or '',  # ubicazione
                r[4] or 1,   # quantita
                r[5] or '',  # stato
                str(r[6]).replace('.', ',') if r[6] else '',  # valore stimato (formato italiano)
                str(r[7]).replace('.', ',') if r[7] else '',  # valore assicurato
                str(r[8]) if r[8] else '',  # data acquisto
                r[9] or '',  # fornitore
                r[10] or '',  # provenienza
                r[11] or '',  # note
            ])

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue().encode('utf-8-sig')]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=inventario_beni.csv"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


# ============================================
# EXPORT EXCEL
# ============================================

@router.get("/export/excel")
async def export_excel(
    current_user: dict = Depends(get_current_user),
    x_ente_id: str = Header(None, alias="X-Ente-Id")
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl non installato. Usa il formato CSV.")

    ente_id = get_ente_id(current_user, x_ente_id)
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT
                b.numero_progressivo, b.descrizione, c.nome, u.nome,
                b.quantita, b.stato_conservazione, b.valore_stimato,
                b.valore_assicurato, b.data_acquisto, b.fornitore,
                b.provenienza, b.note
            FROM beni_inventario b
            LEFT JOIN inventario_categorie c ON b.categoria_id = c.id
            LEFT JOIN inventario_ubicazioni u ON b.ubicazione_id = u.id
            WHERE b.ente_id = %s AND b.stato = 'attivo'
            ORDER BY b.numero_progressivo
        """, (ente_id,))
        rows = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario Beni"

        # Header con stile
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A2E55", end_color="1A2E55", fill_type="solid")

        for col_idx, col_name in enumerate(COLONNE_CSV, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=r[0])
            ws.cell(row=row_idx, column=2, value=r[1] or '')
            ws.cell(row=row_idx, column=3, value=r[2] or '')
            ws.cell(row=row_idx, column=4, value=r[3] or '')
            ws.cell(row=row_idx, column=5, value=r[4] or 1)
            ws.cell(row=row_idx, column=6, value=r[5] or '')
            ws.cell(row=row_idx, column=7, value=float(r[6]) if r[6] else None)
            ws.cell(row=row_idx, column=8, value=float(r[7]) if r[7] else None)
            ws.cell(row=row_idx, column=9, value=str(r[8]) if r[8] else '')
            ws.cell(row=row_idx, column=10, value=r[9] or '')
            ws.cell(row=row_idx, column=11, value=r[10] or '')
            ws.cell(row=row_idx, column=12, value=r[11] or '')

        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventario_beni.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


# ============================================
# TEMPLATE EXCEL PER IMPORTAZIONE
# ============================================

@router.get("/import/template")
async def download_template(
    current_user: dict = Depends(get_current_user)
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl non installato")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Importazione"

    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2E55", end_color="1A2E55", fill_type="solid")

    colonne = ["Descrizione *", "Categoria *", "Ubicazione *", "Quantita",
               "Stato Conservazione", "Valore Stimato", "Data Acquisto",
               "Fornitore", "Provenienza", "Note"]

    for col_idx, col_name in enumerate(colonne, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    # Riga esempio
    ws.cell(row=2, column=1, value="Calice d'argento sec. XVIII")
    ws.cell(row=2, column=2, value="Vasi sacri")
    ws.cell(row=2, column=3, value="Sagrestia")
    ws.cell(row=2, column=4, value=1)
    ws.cell(row=2, column=5, value="buono")
    ws.cell(row=2, column=6, value=500.00)
    ws.cell(row=2, column=7, value="2020-01-15")
    ws.cell(row=2, column=8, value="")
    ws.cell(row=2, column=9, value="Donazione")
    ws.cell(row=2, column=10, value="Restaurato nel 2019")

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_importazione_beni.xlsx"}
    )


# ============================================
# IMPORT DA CSV/EXCEL
# ============================================

@router.post("/import")
async def import_beni(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    x_ente_id: str = Header(None, alias="X-Ente-Id")
):
    ente_id = get_ente_id(current_user, x_ente_id)

    contents = await file.read()
    filename = file.filename.lower()

    rows_data = []

    if filename.endswith('.csv'):
        # Parse CSV
        text = contents.decode('utf-8-sig')
        reader = csv.reader(StringIO(text), delimiter=';')
        headers = next(reader, None)
        if not headers:
            raise HTTPException(status_code=400, detail="File CSV vuoto")
        for row in reader:
            if len(row) < 3:
                continue
            rows_data.append({
                "descrizione": row[0].strip() if len(row) > 0 else "",
                "categoria": row[1].strip() if len(row) > 1 else "",
                "ubicazione": row[2].strip() if len(row) > 2 else "",
                "quantita": row[3].strip() if len(row) > 3 else "1",
                "stato_conservazione": row[4].strip().lower() if len(row) > 4 else "",
                "valore_stimato": row[5].strip().replace(',', '.') if len(row) > 5 else "",
                "data_acquisto": row[6].strip() if len(row) > 6 else "",
                "fornitore": row[7].strip() if len(row) > 7 else "",
                "provenienza": row[8].strip() if len(row) > 8 else "",
                "note": row[9].strip() if len(row) > 9 else "",
            })

    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=501, detail="openpyxl non installato")

        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        next(row_iter, None)  # Skip header

        for row in row_iter:
            if not row or not row[0]:
                continue
            rows_data.append({
                "descrizione": str(row[0]).strip() if row[0] else "",
                "categoria": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "ubicazione": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "quantita": str(row[3]).strip() if len(row) > 3 and row[3] else "1",
                "stato_conservazione": str(row[4]).strip().lower() if len(row) > 4 and row[4] else "",
                "valore_stimato": str(row[5]).strip() if len(row) > 5 and row[5] else "",
                "data_acquisto": str(row[6]).strip() if len(row) > 6 and row[6] else "",
                "fornitore": str(row[7]).strip() if len(row) > 7 and row[7] else "",
                "provenienza": str(row[8]).strip() if len(row) > 8 and row[8] else "",
                "note": str(row[9]).strip() if len(row) > 9 and row[9] else "",
            })
    else:
        raise HTTPException(status_code=400, detail="Formato file non supportato. Usa CSV o Excel (.xlsx)")

    if not rows_data:
        raise HTTPException(status_code=400, detail="Nessun dato trovato nel file")

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Carica categorie e ubicazioni per match per nome
        cur.execute("""
            SELECT id, LOWER(TRIM(nome)) FROM inventario_categorie
            WHERE ente_id = %s AND attivo = TRUE
        """, (ente_id,))
        cat_map = {r[1]: str(r[0]) for r in cur.fetchall()}

        cur.execute("""
            SELECT id, LOWER(TRIM(nome)) FROM inventario_ubicazioni
            WHERE ente_id = %s AND attivo = TRUE
        """, (ente_id,))
        ub_map = {r[1]: str(r[0]) for r in cur.fetchall()}

        # Prossimo numero progressivo
        cur.execute("""
            SELECT COALESCE(MAX(numero_progressivo), 0) FROM beni_inventario WHERE ente_id = %s
        """, (ente_id,))
        next_num = cur.fetchone()[0] + 1

        importati = 0
        errori = []

        for idx, row in enumerate(rows_data, 2):
            descrizione = row["descrizione"]
            if not descrizione:
                errori.append(f"Riga {idx}: descrizione mancante")
                continue

            cat_nome = row["categoria"].lower().strip()
            ub_nome = row["ubicazione"].lower().strip()

            cat_id = cat_map.get(cat_nome)
            ub_id = ub_map.get(ub_nome)

            if not cat_id and cat_nome:
                errori.append(f"Riga {idx}: categoria '{row['categoria']}' non trovata")
                continue
            if not ub_id and ub_nome:
                errori.append(f"Riga {idx}: ubicazione '{row['ubicazione']}' non trovata")
                continue

            # Parse valori
            try:
                quantita = int(row["quantita"]) if row["quantita"] else 1
            except ValueError:
                quantita = 1

            valore = None
            if row["valore_stimato"]:
                try:
                    valore = float(row["valore_stimato"])
                except ValueError:
                    pass

            data_acquisto = row["data_acquisto"] if row["data_acquisto"] else None
            stato_cons = row["stato_conservazione"] if row["stato_conservazione"] in ('ottimo', 'buono', 'discreto', 'restauro', 'scadente') else None

            bene_id = str(uuid.uuid4())
            cur.execute("""
                INSERT INTO beni_inventario (
                    id, ente_id, numero_progressivo, categoria_id, ubicazione_id,
                    descrizione, quantita, stato_conservazione, valore_stimato,
                    data_acquisto, fornitore, provenienza, note,
                    created_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                bene_id, ente_id, next_num, cat_id, ub_id,
                descrizione, quantita, stato_cons, valore,
                data_acquisto, row["fornitore"] or None, row["provenienza"] or None, row["note"] or None,
                current_user['user_id']
            ))
            next_num += 1
            importati += 1

        conn.commit()

        result = {"importati": importati, "totale_righe": len(rows_data)}
        if errori:
            result["errori"] = errori[:20]  # Max 20 errori
        return result

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()
